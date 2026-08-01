"""Unit test endpoint ekspor laporan (Phase 36).

Jalankan dari folder backend/ (venv aktif):

    python test_report_api.py

Aman: database SQLite in-memory (PostgreSQL tidak disentuh),
TensorFlow di-stub bila tidak terpasang. Butuh openpyxl
(pip install openpyxl).
"""

import csv
import io
import sys
import types
from datetime import datetime, timedelta, timezone

# --- Stub TensorFlow bila tidak terpasang ---
try:  # noqa: SIM105
    import tensorflow  # noqa: F401
except ImportError:
    tf_stub = types.ModuleType("tensorflow")
    tf_stub.lite = types.SimpleNamespace(Interpreter=object)
    tf_stub.__version__ = "stub"
    sys.modules["tensorflow"] = tf_stub

from fastapi.testclient import TestClient  # noqa: E402
from openpyxl import load_workbook  # noqa: E402
from sqlalchemy import create_engine  # noqa: E402
from sqlalchemy.orm import sessionmaker  # noqa: E402
from sqlalchemy.pool import StaticPool  # noqa: E402

from app.core.security import hash_password  # noqa: E402
from app.db.base import Base  # noqa: E402
import app.models  # noqa: E402,F401
from app.models import DetectionResult, User  # noqa: E402
from app.db.session import get_db  # noqa: E402
from app.main import app  # noqa: E402

PASS = 0
FAIL = 0


def check(name: str, condition: bool, detail: str = "") -> None:
    """Mencetak hasil satu pengecekan dan mencatat statistik."""
    global PASS, FAIL  # noqa: PLW0603
    ok = "LOLOS" if condition else "GAGAL"
    if condition:
        PASS += 1
    else:
        FAIL += 1
    print(f"  [{ok}] {name}" + (f" | {detail}" if detail else ""))


def main() -> int:
    engine = create_engine(
        "sqlite://",
        future=True,
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    Base.metadata.create_all(bind=engine)
    TestSession = sessionmaker(bind=engine, future=True)

    def override_get_db():
        session = TestSession()
        try:
            yield session
        finally:
            session.close()

    app.dependency_overrides[get_db] = override_get_db
    client = TestClient(app)

    # Seed: admin + 5 record (3 Tunas Air baru, 1 Buah lama, 1 NoDet).
    now = datetime.now(timezone.utc)
    seed = [
        ("Tunas Air", 0),
        ("Tunas Air", 1),
        ("Tunas Air", 2),
        ("Buah Siap", 30),
        ("No Detection", 31),
    ]
    db = TestSession()
    db.add(
        User(
            username="admin_test",
            email="admin_test@melon.local",
            password_hash=hash_password("admin123"),
            full_name="Admin Test",
        )
    )
    for i, (label, days_ago) in enumerate(seed):
        db.add(
            DetectionResult(
                image_path=f"uploads/report_{i}.jpg",
                label=label,
                action="None",
                detected_at=now - timedelta(days=days_ago),
            )
        )
    db.commit()
    db.close()

    token = client.post(
        "/auth/login",
        json={"username": "admin_test", "password": "admin123"},
    ).json()["access_token"]
    headers = {"Authorization": f"Bearer {token}"}

    # ============ [1] Proteksi autentikasi ============
    print("\n[1] Proteksi autentikasi")
    check("CSV tanpa token -> 401",
          client.get("/report/export/csv").status_code == 401)
    check("Excel tanpa token -> 401",
          client.get("/report/export/excel").status_code == 401)

    # ============ [2] Export CSV ============
    print("\n[2] Export CSV")
    r = client.get("/report/export/csv", headers=headers)
    check("status 200", r.status_code == 200)
    check("content-type text/csv",
          r.headers["content-type"].startswith("text/csv"))
    check(
        "Content-Disposition attachment .csv",
        "attachment" in r.headers.get("content-disposition", "")
        and ".csv" in r.headers.get("content-disposition", ""),
    )
    rows = list(csv.reader(io.StringIO(r.text)))
    check("header 6 kolom", rows[0] == [
        "id", "label", "action", "image_path",
        "annotated_image_path", "detected_at",
    ])
    check("5 baris data", len(rows) - 1 == 5, f"data={len(rows) - 1}")

    # ============ [3] Filter pada CSV ============
    print("\n[3] Filter pada CSV")
    r = client.get("/report/export/csv?label=Tunas%20Air", headers=headers)
    rows = list(csv.reader(io.StringIO(r.text)))
    check("filter label -> 3 baris", len(rows) - 1 == 3)
    check("semua baris berlabel Tunas Air",
          all(row[1] == "Tunas Air" for row in rows[1:]))

    date_from = (now - timedelta(days=5)).date().isoformat()
    r = client.get(
        f"/report/export/csv?date_from={date_from}", headers=headers
    )
    rows = list(csv.reader(io.StringIO(r.text)))
    check("filter date_from -> 3 baris", len(rows) - 1 == 3)

    date_to = (now - timedelta(days=10)).date().isoformat()
    r = client.get(
        f"/report/export/csv?date_to={date_to}", headers=headers
    )
    rows = list(csv.reader(io.StringIO(r.text)))
    check("filter date_to -> 2 baris", len(rows) - 1 == 2)

    # ============ [4] Export Excel ============
    print("\n[4] Export Excel")
    r = client.get("/report/export/excel", headers=headers)
    check("status 200", r.status_code == 200)
    check(
        "content-type xlsx",
        "spreadsheetml" in r.headers["content-type"],
    )
    workbook = load_workbook(io.BytesIO(r.content), read_only=True)
    sheet = workbook.active
    xl_rows = list(sheet.iter_rows(values_only=True))
    check("sheet 'Riwayat Deteksi'", sheet.title == "Riwayat Deteksi")
    check("header 6 kolom", list(xl_rows[0]) == list((
        "id", "label", "action", "image_path",
        "annotated_image_path", "detected_at",
    )))
    check("5 baris data", len(xl_rows) - 1 == 5)

    r = client.get(
        "/report/export/excel?label=Buah%20Siap", headers=headers
    )
    workbook = load_workbook(io.BytesIO(r.content), read_only=True)
    xl_rows = list(workbook.active.iter_rows(values_only=True))
    check("filter label Excel -> 1 baris", len(xl_rows) - 1 == 1)

    print(f"\n=== HASIL: {PASS} lolos, {FAIL} gagal ===")
    return 0 if FAIL == 0 else 1


if __name__ == "__main__":
    sys.exit(main())